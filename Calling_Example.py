# Excel文件的修改
import OfficeHelper
import os
from openpyxl.styles import Alignment,Font,Border,Side,PatternFill

base_dir1='C:/Users/NUC/Desktop/excels'
for first_floor in os.listdir(base_dir1):
    file_dir=base_dir1+'/'+first_floor
    file_suffix=file_dir.split('.')[-1]
    if file_suffix == file_dir.split('.')[-1]=='xlsx':
        excel=OfficeHelper.ExcelFormatHelper(file_dir)
        excel.change_cell_format('Test_Sheet',1,1,Font(name='微软雅黑',size=14,bold=True,color='FF0000'))
        excel.change_cell_fill_format('Test_Sheet',1,1,PatternFill(fill_type='solid',fgColor='FF0000'))
        excel.change_cell_alignment('Test_Sheet',1,1,Alignment(horizontal='center',vertical='center'))
        excel.change_cell_border('Test_Sheet',1,1,Border(left=Side(border_style='thin',color='FF0000'),
                                                            right=Side(border_style='thin',color='FF0000'),
                                                            top=Side(border_style='thin',color='FF0000'),
                                                            bottom=Side(border_style='thin',color='FF0000')))
        excel.write_to_cell('Test_Sheet',26,26,'测试')
        excel.delete_cell('Test_Sheet',1,1)
        excel.get_cell_location('Test_Sheet','测试')
        excel.auto_fit_all_columns()
        excel.set_row_height('Test_Sheet',19,14.4)
        excel.merge_cells('Test_Sheet',3,1,3,2)
        excel.merge_area_cells('Test_Sheet',3,1,3,2)
        excel.new_sheet('Test_Sheet1')
        excel.change_sheet_name('Test_Sheet1','Test_Sheet2')
        excel.delete_sheet('Test_Sheet2')
        excel.change_number_format('Test_Sheet','0.00')
        excel.save_excel()
        excel.close_excel()

# Word文件的修改
import os
import OfficeHelper 
from docx.shared import Pt
from docx.enum.text import WD_ALIGN_PARAGRAPH

base_dir2='C:/Users/NUC/Desktop/words'
for first_floor in os.listdir(base_dir2):
    file_dir=base_dir2+'/'+first_floor
    file_suffix=file_dir.split('.')[-1]
    if file_suffix == file_dir.split('.')[-1]=='docx':
        word=OfficeHelper.WordFormatHelper(file_dir)
        word.change_format(u'Times New Roman',12,[0,0,255])
        word.change_align(WD_ALIGN_PARAGRAPH.LEFT)
        word.change_line_spacing(1)
        word.change_paragraph_spacing(Pt(0),Pt(0))
        word.replace_all('hello','world')
        word.indent_first_line()
        word.save_word(file_dir)
